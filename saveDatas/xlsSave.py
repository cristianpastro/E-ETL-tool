#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Sep 07 18:55:03 2020

@author: cristian
"""

import xlsxwriter
from openpyxl import load_workbook


def cria_xlsx(file_name):
    """
     Cria um arquivo vazio no formato .xslx

     Parameters
     ----------
     file_name : str
         Nome do arquivo que será salvo (sem o .xlsx no final)

     """

    workbook = xlsxwriter.Workbook(file_name + '.xlsx')
    worksheet = workbook.add_worksheet()
    workbook.close()


def add_set_to_xlsx(file_xlsx, set_datas):
    """
     A partir de um conjunto de dados, salva os dados em um arquivo .xlsx

     Parameters
     ----------
     file_name : str
         Nome do arquivo que será salvo (sem o .xlsx no final)
     set_datas : list of lists
         Lista de listas com os dados a serem salvos. Cada lista interna contém uma linha que
         será salvo no arquivo .xlsx

     """

    cria_xlsx(file_xlsx)

    row_number = 1
    wb = load_workbook(file_xlsx + '.xlsx')
    ws = wb.get_sheet_by_name("Sheet1")

    for line in set_datas:
        col = 1
        for item in line:
            c = ws.cell(row=row_number, column=col)
            c.value = item
            col += 1
        row_number += 1

    wb.save(file_xlsx + '.xlsx')
